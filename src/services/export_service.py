"""学生数据导出 Service（v8.1 - 通用列树 + 行展开自动推断 + 不持久化）

数据流：
    1. 接收 ExportUsersRequest（含 columns 树、filters）
    2. 校验列树（白名单 + 层级约束 + 字段完整性）
    3. 查学生（按 filters + studentIds/excludedIds）
    4. 预加载 applications（按 category_id 收集 PASSED 列表）+ extra_info_field（name → id）
    5. 自动推断每个 category 的行展开（看子列含 application_*）
    6. openpyxl 生成多级表头（DFS 算 col_start/col_end + row_start/row_end）+ 数据行
    7. 流式返回 xlsx 文件

详见 docs/docs-backend/导出表格/export-后端实现方案.md
"""
from __future__ import annotations

import io
import logging
from collections import defaultdict
from decimal import Decimal
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.app.schemas.errors import BadRequestError
from src.app.schemas.export import (
    APPLICATION_APPLY,
    APPLICATION_ATTR,
    APPLICATION_FIELDS,
    APPLICATION_FIELD,
    APPLICATION_GAIN,
    APPLICATION_REMARK,
    APPLICATION_STATUS,
    APPLICATION_WEIGHTED_SUM,
    CATEGORY,
    CONSTRAINED_SOURCES,
    CUSTOM,
    ExportColumnNode,
    ExportUsersRequest,
    USER_BASIC,
    USER_BASIC_FIELDS,
    USER_EXTRA,
)
from src.models.application import Application, ApplicationStatus
from src.models.extra_info_field import ExtraInfoField
from src.models.user import User

logger = logging.getLogger(__name__)


# ════════════════════════════════════════════════════════════════════════
# 节点坐标分配（DFS）
# ════════════════════════════════════════════════════════════════════════

class NodeCoord:
    """列树节点的 Excel 坐标（后端运行时计算，前端不感知）"""  # 记录节点在 Excel 中的行列范围
    __slots__ = ("node", "col_start", "col_end", "row_start", "row_end")

    def __init__(self, node: ExportColumnNode) -> None:
        self.node = node
        self.col_start = 0
        self.col_end = 0
        self.row_start = 0
        self.row_end = 0


def _assign_coordinates(columns: List[ExportColumnNode]) -> List[NodeCoord]:
    """DFS 给每个节点分配 col_start/col_end + row_start/row_end。

    col 范围：后序遍历，叶子 [start,start]；父 [first.start, last.end]
    row 范围：前序遍历，父 row=r，子 row=r+1；父 row_end=r（父只占 1 行）
    """  # 计算每个列树节点在 Excel 中的占位范围（横向列区间 + 纵向行区间）
    # 先建全树 coord 映射（确保子节点在递归时已注册）
    coords_map: Dict[str, NodeCoord] = {}

    def get_or_create(node: ExportColumnNode) -> NodeCoord:
        """按节点 id 获取或新建一个 NodeCoord（DFS 期间惰性创建子节点 coord）"""  # 获取或惰性创建节点的 coord 容器
        if node.id not in coords_map:
            coords_map[node.id] = NodeCoord(node)
        return coords_map[node.id]

    # 先注册全部节点
    def register_all(nodes: List[ExportColumnNode]) -> None:
        """递归注册整棵子树所有节点到 coords_map（先建索引再做分配）"""  # 预先把整棵树的节点都登记到 coords_map
        for n in nodes:
            get_or_create(n)
            if n.children:
                register_all(n.children)

    register_all(columns)
    # ══════ FIX: 顶级 cols 按 sortOrder 排序 ══════
    # 前端拖拽后可能存在"数组顺序 = DOM 顺序"但 sortOrder 与数组顺序不一致的情况，
    # 必须显式按 sortOrder 排序保证正确性（dfs_col 内部对 children 已排序）
    sorted_columns = sorted(columns, key=lambda n: n.sortOrder)
    coords = [get_or_create(n) for n in sorted_columns]

    # DFS col 范围（后序）
    def dfs_col(coord: NodeCoord, start_col: int) -> int:
        """DFS 计算 col 范围（后序），返回下一个可用 col 游标"""  # 自底向上算出当前节点及其子树的列起止
        children_sorted = sorted(coord.node.children, key=lambda c: c.sortOrder)
        if not children_sorted:
            coord.col_start = start_col
            coord.col_end = start_col
            return start_col + 1
        cur = start_col
        for child_node in children_sorted:
            child_coord = get_or_create(child_node)
            cur = dfs_col(child_coord, cur)
        coord.col_start = start_col
        coord.col_end = cur - 1
        return cur

    # DFS row 范围（前序）：父 row=r，子 row=r+1；父 row_end=r（占 1 行）
    def dfs_row(coord: NodeCoord, row: int) -> int:
        """DFS 计算 row 范围（前序），返回当前子树最终行下标"""  # 自顶向下算出当前节点及其子树的行起止
        coord.row_start = row
        children_sorted = sorted(coord.node.children, key=lambda c: c.sortOrder)
        if not children_sorted:
            coord.row_end = row
            return row + 1
        # 所有子节点都在同一行 row+1；递归取最深的 row_end 作为父的"占用结束行"
        max_child_end = row
        for child_node in children_sorted:
            child_coord = get_or_create(child_node)
            child_end = dfs_row(child_coord, row + 1)
            max_child_end = max(max_child_end, child_end)
        coord.row_end = row  # 父节点只占 1 行
        return max_child_end + 1  # cursor 推进到最深子树后

    # 第一遍：col 分配（按顶级顺序累加 col cursor）
    col_cursor = 0
    for coord in coords:
        col_cursor = dfs_col(coord, col_cursor)

    # 第二遍：row 分配（每个顶级节点独立从 row=0 开始 —— 顶级都在第 1 行）
    for coord in coords:
        dfs_row(coord, 0)

    # ★ 修复：返回 coords_map.values() 而不是 coords
    # 因为 dfs_col / dfs_row 会把子节点注册到 coords_map，但 coords 只包含顶级
    return list(coords_map.values())


# ════════════════════════════════════════════════════════════════════════
# 列树校验
# ════════════════════════════════════════════════════════════════════════

def _validate_columns(columns: List[ExportColumnNode]) -> None:
    """递归校验列树合法性（白名单 + 层级约束 + 字段完整性）。

    抛出 BadRequestError 时直接返回 400 + 错误信息。
    """  # 校验列树合法性（白名单 + 层级约束 + 必填字段），不合规直接 400
    seen_ids: Set[str] = set()
    valid_sources = set(CONSTRAINED_SOURCES) | {USER_BASIC, USER_EXTRA, CATEGORY, CUSTOM}

    def walk(node: ExportColumnNode, parent: Optional[ExportColumnNode]) -> None:
        """DFS 递归校验每个节点及其子树"""  # DFS 校验单个节点并递归处理子树
        # 1. source 白名单（理论上 Pydantic Literal 已拦截，这里防御性再校验）
        if node.source not in valid_sources:
            raise BadRequestError(
                f"列 '{node.label}' 的 source='{node.source}' 不在白名单内"
            )

        # 2. id 唯一
        if node.id in seen_ids:
            raise BadRequestError(f"列树节点 id 重复: {node.id}")
        seen_ids.add(node.id)

        # 3. parentId 指向合法父节点
        expected_parent_id = parent.id if parent else None
        if node.parentId != expected_parent_id:
            raise BadRequestError(
                f"节点 '{node.label}' 的 parentId={node.parentId} "
                f"与实际父 '{expected_parent_id}' 不一致"
            )

        # 4. ★ 层级约束：application_* 必须挂 category 下
        if node.source in CONSTRAINED_SOURCES:
            if parent is None or parent.source != CATEGORY:
                parent_label = parent.label if parent else "顶级"
                parent_source = parent.source if parent else "(none)"
                raise BadRequestError(
                    f"列 '{node.label}' (source={node.source}) 必须挂在 category 节点下，"
                    f"当前父节点是 '{parent_label}' (source={parent_source})"
                )

        # 5. application_attr 必须有 ruleName
        if node.source == APPLICATION_ATTR and not node.ruleName:
            raise BadRequestError(
                f"application_attr 列 '{node.label}' 必须指定 ruleName"
            )

        # 5b. application_field 必须有 appField 且在白名单内
        if node.source == APPLICATION_FIELD:
            if not node.appField:
                raise BadRequestError(
                    f"application_field 列 '{node.label}' 必须指定 appField"
                )
            if node.appField not in APPLICATION_FIELDS:
                raise BadRequestError(
                    f"application_field 列 '{node.label}' 的 appField='{node.appField}' "
                    f"不在白名单 {sorted(APPLICATION_FIELDS)} 内"
                )

        # 6. user_extra 必须有 fieldPath
        if node.source == USER_EXTRA and not node.fieldPath:
            raise BadRequestError(
                f"user_extra 列 '{node.label}' 必须指定 fieldPath"
            )

        # 7. user_basic 必须有 basicField（且必须在白名单内）
        if node.source == USER_BASIC:
            if not node.basicField:
                raise BadRequestError(
                    f"user_basic 列 '{node.label}' 必须指定 basicField"
                )
            if node.basicField not in USER_BASIC_FIELDS:
                raise BadRequestError(
                    f"user_basic 列 '{node.label}' 的 basicField='{node.basicField}' "
                    f"不在白名单 {sorted(USER_BASIC_FIELDS)} 内"
                )

        # 8. category 必须有 categoryId
        if node.source == CATEGORY and node.categoryId is None:
            raise BadRequestError(
                f"category 节点 '{node.label}' 必须指定 categoryId"
            )

        # 9. category 节点本身不应有 user_* 字段（防御）
        if node.source == CATEGORY:
            if node.basicField or node.fieldPath:
                raise BadRequestError(
                    f"category 节点 '{node.label}' 不应携带 basicField / fieldPath"
                )

        # 9b. custom 节点必须是容器（必须有至少 1 个子列），否则成为空列
        if node.source == CUSTOM:
            if not node.children:
                raise BadRequestError(
                    f"自定义根列 '{node.label}' 必须至少包含 1 个子列（请在编辑弹窗中添加子列后再导出）"
                )

        # 10. 递归子列
        for child in node.children:
            walk(child, node)

    for node in columns:
        walk(node, None)


# ════════════════════════════════════════════════════════════════════════
# 行展开自动推断
# ════════════════════════════════════════════════════════════════════════

def _walk_descendants(node: ExportColumnNode):
    """DFS 遍历 node 全部后代（含 node 自己）"""  # 递归产出某节点及其所有后代
    yield node
    for child in node.children:
        yield from _walk_descendants(child)


def _category_has_application_columns(cat_node: ExportColumnNode) -> bool:
    """判断 category 节点的子树里是否含 application_* 字段"""  # 判断某 category 子树是否需要按 application 展开行
    for n in _walk_descendants(cat_node):
        if n.source in CONSTRAINED_SOURCES:
            return True
    return False


def _infer_max_app_count(
    category_nodes: List[ExportColumnNode],
    app_cache: Dict[int, Dict[int, List[Application]]],
    user_id: int,
    max_per_cat: int,
) -> int:
    """对单个学生，返回所有"按 application 展开"的 category 中 PASSED application 数的最大值"""  # 计算单个学生在所有 category 中需要展开的最大 application 行数
    max_count = 0
    for cat in category_nodes:
        if not _category_has_application_columns(cat):
            continue
        cat_id = cat.categoryId
        if cat_id is None:
            continue
        apps = app_cache.get(user_id, {}).get(cat_id, [])
        max_count = max(max_count, min(len(apps), max_per_cat))
    return max_count


# ════════════════════════════════════════════════════════════════════════
# 取值逻辑
# ════════════════════════════════════════════════════════════════════════

def _resolve_user_basic_value(user: User, basic_field: str) -> Any:
    """user_basic 列取值"""  # 按 basicField 白名单从 User 模型取基础字段值
    # studentId 特殊处理：优先 users.student_id，fallback 到 extract_student_id(username)
    if basic_field == "studentId":
        if user.student_id:
            return user.student_id
        return User.extract_student_id(user.username) or ""
    if basic_field == "fullName":
        return user.full_name or ""
    if basic_field == "phone":
        return user.phone or ""
    if basic_field == "department":
        return user.department or ""
    if basic_field == "major":
        return user.major or ""
    if basic_field == "grade":
        return user.grade if user.grade is not None else ""
    if basic_field == "enrollmentYear":
        return user.enrollment_year if user.enrollment_year is not None else ""
    if basic_field == "graduationYear":
        return user.graduation_year if user.graduation_year is not None else ""
    if basic_field == "gender":
        return user.gender or ""
    if basic_field == "idCardNumber":
        return user.id_card_number or ""
    if basic_field == "username":
        return user.username or ""
    if basic_field == "lastLoginAt":
        return user.last_login_at or ""
    return ""


def _transform_grade(value: Any) -> Any:
    """grade 1/2/3/4 → "大一"/"大二"/"大三"/"大四"（≤4）"""  # 把年级数字转中文显示文本
    if isinstance(value, int) and 1 <= value <= 4:
        return {1: "大一", 2: "大二", 3: "大三", 4: "大四"}.get(value, value)
    return value


def _resolve_user_extra_value(
    user: User, field_path: str, extra_field_id: int
) -> Any:
    """user_extra 列取值（extra_info JSON 中 f_{field_id} 取值）"""  # 按 fieldPath 对应 id 从 user.extra_info 读取扩展字段
    extra = user.extra_info or {}
    key = f"f_{extra_field_id}"
    return extra.get(key, "")


def _resolve_app_value(app: Application, col: ExportColumnNode) -> Any:
    """application_* 列取值（按 row_idx 已定位到 app）"""  # 按列 source 类型从已定位到的 application 取出对应字段值
    if col.source == APPLICATION_APPLY:
        return float(app.apply_score) if app.apply_score is not None else ""
    if col.source == APPLICATION_GAIN:
        return float(app.gain_score) if app.gain_score is not None else ""
    if col.source == APPLICATION_ATTR:
        rule_info = app.rule_info or {}
        if col.ruleName is None:
            return ""
        return rule_info.get(col.ruleName, "")
    if col.source == APPLICATION_STATUS:
        return app.status or ""
    if col.source == APPLICATION_REMARK:
        # Application ORM 无 remark 字段（remark 在 ApplicationOperation 表），
        # v8.1 暂时返回空；未来如需可从 ApplicationOperation 拼装最近一条 remark。
        return ""
    if col.source == APPLICATION_FIELD:
        # appField 指向 Application ORM 的任意白名单字段
        return _resolve_application_field(app, col.appField)
    return ""


def _resolve_weighted_sum_formula(
    col: ExportColumnNode,
    coords: List[NodeCoord],
    start_row: int,
    row_count: int,
) -> Optional[str]:
    """构造 application_weighted_sum 列在某个学生的公式

    返回形如 "=IFERROR(0.3*B2:B5,0) + IFERROR(0.7*C2:C5,0)" 的字符串（带等号）。

    每一项都包裹 IFERROR(..,0)，原因：
    - 加权列可能引用 application_attr（属性，文本）/ application_status（状态，文本）/
      application_remark（备注，文本）。这些子列的值不是数字，Excel 在
      `weight * 文本` 时会返回 #VALUE!，并污染整个 `+` 表达式。
    - 用 IFERROR 把每一项独立兜底为 0，整列仍可正常求和（且不会显示 #VALUE!）。
    - 如果用户选的都是 application_apply / application_gain / application_field(数字)，
      IFERROR 不会触发，正常计算。

    Args:
        col: 当前 weighted_sum 列节点（含 weightedColumnIds + weightedWeights）
        coords: 当前渲染范围内所有 coord（用于按 id 找 sibling）
        start_row: 当前学生在 Excel 中的起始行（0-based）
        row_count: 当前学生展开的总行数（max_count）
    """  # 生成 application_weighted_sum 列在某学生展开区间的 Excel 公式（IFERROR 兜底）
    if not col.weightedColumnIds or not col.weightedWeights:
        return None
    if len(col.weightedColumnIds) != len(col.weightedWeights):
        logger.warning(
            "weighted_sum 列 '%s' 的 weightedColumnIds (%d) 和 weightedWeights (%d) 长度不一致",
            col.label,
            len(col.weightedColumnIds),
            len(col.weightedWeights),
        )
        return None

    from openpyxl.utils import get_column_letter

    # 按 id 找 sibling coord 的快速查找
    coords_by_id = {c.node.id: c for c in coords}

    parts: List[str] = []
    for child_id, weight in zip(col.weightedColumnIds, col.weightedWeights):
        sibling_coord = coords_by_id.get(child_id)
        if not sibling_coord:
            logger.warning(
                "weighted_sum 列 '%s' 引用了不存在的 sibling id: %s",
                col.label,
                child_id,
            )
            continue
        col_letter = get_column_letter(sibling_coord.col_start + 1)
        first_row = start_row + 1  # openpyxl 1-based
        last_row = start_row + row_count
        # 每项用 IFERROR(..,0) 独立兜底，避免任一非数字项污染整列
        parts.append(
            f"IFERROR({weight}*{col_letter}{first_row}:{col_letter}{last_row},0)"
        )

    if not parts:
        return None
    # 整体再用 IFERROR(..,"") 兜底：
    # - 当整行所有项都为空/无效时（如某学生没有任何 application 数据，
    #   所有 IFERROR 都返回 0），结果会是 0——这里改成 "" 显示空。
    # - 任一正常数字项仍正常求和（IFERROR 不触发）。
    return f"=IFERROR({' + '.join(parts)},\"\")"


def _resolve_application_field(app: Application, app_field: Optional[str]) -> Any:
    """application_field 列取值（按 appField 指向 Application ORM 字段）"""  # 通用字段取值：按白名单内的 ORM 字段名读 application 属性
    if not app_field:
        return ""
    value = getattr(app, app_field, None)
    if value is None:
        return ""
    # datetime → ISO 字符串
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat(sep=" ", timespec="seconds")
        except TypeError:
            return value.isoformat()
    # Decimal → float
    if hasattr(value, "__class__") and value.__class__.__name__ == "Decimal":
        try:
            return float(value)
        except (TypeError, ValueError):
            return str(value)
    return value


# ════════════════════════════════════════════════════════════════════════
# 数据查询
# ════════════════════════════════════════════════════════════════════════

async def _query_students(
    db: AsyncSession,
    filters: Any,
    student_ids: Optional[List[int]],
    excluded_ids: Optional[List[int]],
) -> List[User]:
    """按过滤条件查学生（admin 用）"""  # 按 filters + studentIds/excludedIds 查学生列表（按 id 升序）
    stmt = select(User)

    if filters.username:
        stmt = stmt.where(User.username.ilike(f"%{filters.username}%"))
    if filters.fullName:
        stmt = stmt.where(User.full_name.ilike(f"%{filters.fullName}%"))
    if filters.major:
        stmt = stmt.where(User.major.ilike(f"%{filters.major}%"))
    if filters.department:
        stmt = stmt.where(User.department.ilike(f"%{filters.department}%"))
    if filters.grade is not None:
        stmt = stmt.where(User.grade == filters.grade)
    if filters.enrollmentYear is not None:
        stmt = stmt.where(User.enrollment_year == filters.enrollmentYear)
    if filters.graduationYear is not None:
        stmt = stmt.where(User.graduation_year == filters.graduationYear)

    if student_ids:
        stmt = stmt.where(User.id.in_(student_ids))
    if excluded_ids:
        stmt = stmt.where(User.id.notin_(excluded_ids))

    # 按 id ASC 保持稳定顺序（不依赖 created_at，避免同 created_at 的乱序）
    stmt = stmt.order_by(User.id.asc())

    result = await db.execute(stmt)
    return list(result.scalars().all())


async def _preload_applications(
    db: AsyncSession,
    user_ids: List[int],
    category_ids: List[int],
) -> Dict[int, Dict[int, List[Application]]]:
    """预加载所有相关 application：app_cache[user_id][category_id] = [PASSED apps, ...]

    注意：category_id 直接来自 application.category_id（外键到 template_category.id），
    而非 application.template.category_id。Application.category_id 字段已直接持有分类 id。
    """  # 一次性预加载所有相关学生的 PASSED application，构建嵌套缓存 app_cache[user_id][category_id]
    if not user_ids or not category_ids:
        return {}

    stmt = (
        select(Application)
        .where(Application.user_id.in_(user_ids))
        .where(Application.category_id.in_(category_ids))
        .where(Application.status == ApplicationStatus.PASSED.value)
        .order_by(Application.user_id.asc(), Application.category_id.asc(), Application.id.asc())
    )
    result = await db.execute(stmt)
    apps = list(result.scalars().all())

    cache: Dict[int, Dict[int, List[Application]]] = defaultdict(lambda: defaultdict(list))
    for app in apps:
        cat_id = app.category_id
        if cat_id is None:
            continue
        cache[app.user_id][cat_id].append(app)
    return cache


async def _preload_extra_field_map(
    db: AsyncSession,
    field_paths: List[str],
) -> Dict[str, int]:
    """把 extra_info_field.name → id，建立映射

    若有同名（理论上不应有，业务侧保证），取最小 id。
    """  # 预加载扩展字段 name → id 映射（同 name 取最小 id）
    if not field_paths:
        return {}
    stmt = select(ExtraInfoField).where(ExtraInfoField.name.in_(field_paths))
    result = await db.execute(stmt)
    fields = list(result.scalars().all())
    # 防御性：name 重名时取最小 id（前端按 name 引用，要求唯一）
    by_name: Dict[str, int] = {}
    seen_names: Dict[str, List[int]] = defaultdict(list)
    for f in fields:
        seen_names[f.name].append(f.id)
    for name, ids in seen_names.items():
        by_name[name] = min(ids)
        if len(ids) > 1:
            logger.warning(
                f"extra_info_field.name='{name}' 存在多个 id {ids}，取最小 {min(ids)}"
            )
    return by_name


# ════════════════════════════════════════════════════════════════════════
# 列树遍历辅助
# ════════════════════════════════════════════════════════════════════════

def _collect_category_nodes(columns: List[ExportColumnNode]) -> List[ExportColumnNode]:
    """收集列树里所有 category 节点（递归）"""  # 递归收集列树中所有 category 节点
    result: List[ExportColumnNode] = []
    for col in columns:
        if col.source == CATEGORY:
            result.append(col)
        for child in col.children:
            result.extend(_collect_category_nodes([child]))
    return result


def _collect_field_paths(columns: List[ExportColumnNode]) -> List[str]:
    """收集所有 user_extra 列的 fieldPath（用于预加载）"""  # 收集所有 user_extra 列的 fieldPath 并去重
    paths: List[str] = []
    for col in _walk_all(columns):
        if col.source == USER_EXTRA and col.fieldPath:
            paths.append(col.fieldPath)
    return list(set(paths))  # 去重


def _collect_category_ids(category_nodes: List[ExportColumnNode]) -> List[int]:
    """收集所有 category 节点的 categoryId（用于预加载 application）"""  # 提取所有 category 节点的非空 categoryId 列表
    return [n.categoryId for n in category_nodes if n.categoryId is not None]


def _walk_all(columns: List[ExportColumnNode]):
    """遍历全部节点（含顶级和深层子列）"""  # DFS 产出列树中所有节点（含顶级）
    for col in columns:
        yield from _walk_descendants(col)


def _build_by_id_map(columns: List[ExportColumnNode]) -> Dict[str, ExportColumnNode]:
    """id → node 映射（用于祖先查找）"""  # 把列树扁平化成 id → 节点的字典，便于 O(1) 祖先查找
    return {n.id: n for n in _walk_all(columns)}


def _find_ancestor_category_id_in_map(
    col: ExportColumnNode,
    by_id: Dict[str, ExportColumnNode],
) -> Optional[int]:
    """向上找最近的 category 祖先，返回 categoryId（找不到返回 None）"""  # 在 id 映射中沿 parentId 向上找最近的 category 祖先
    node = col
    while node is not None:
        if node.source == CATEGORY and node.categoryId is not None:
            return node.categoryId
        parent_id = node.parentId
        if parent_id is None:
            return None
        node = by_id.get(parent_id)
    return None


# ════════════════════════════════════════════════════════════════════════
# Excel 渲染
# ════════════════════════════════════════════════════════════════════════

# 表头样式：粗体 + 居中，无填充色（避免遮挡表格）
_HEADER_FONT = Font(bold=False)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _render_header(ws: Worksheet, coords: List[NodeCoord]) -> int:
    """渲染多级表头 + merge_cells。返回表头总行数。"""  # 渲染多级表头并按 coord 范围合并单元格，返回表头总行数
    if not coords:
        return 0

    max_row = max(c.row_end for c in coords)

    # 1. 填标签（只填左上角；非叶子的 merge 由后续步骤完成）
    for coord in coords:
        cell = ws.cell(
            row=coord.row_start + 1,  # openpyxl 1-based
            column=coord.col_start + 1,
            value=coord.node.label,
        )
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN

    # 2. 合并非叶子节点（横向 + 父节点纵向）
    for coord in coords:
        is_leaf = not coord.node.children
        if is_leaf:
            continue
        # 横向合并（多个子列 → 1 个父标签）
        if coord.col_end > coord.col_start:
            ws.merge_cells(
                start_row=coord.row_start + 1,
                end_row=coord.row_end + 1,
                start_column=coord.col_start + 1,
                end_column=coord.col_end + 1,
            )
        # 父节点本身已填值，不需要再纵向合并

    return max_row + 1


def _render_data(
    ws: Worksheet,
    start_row: int,
    user: User,
    coords: List[NodeCoord],
    app_cache: Dict[int, Dict[int, List[Application]]],
    extra_field_map: Dict[str, int],
    by_id: Dict[str, ExportColumnNode],
    max_per_cat: int,
) -> int:
    """渲染单个学生的数据行。

    Returns:
        下一个可用的 start_row（即 start_row + max_count），
        让外层循环能在多个 user 间正确推进，避免互相覆盖。
    """  # 渲染单个学生的数据行（含 application 展开 + 重复列合并），返回下一个可用行号
    # 提取所有 category 节点
    category_nodes: List[ExportColumnNode] = []
    for coord in coords:
        if coord.node.source == CATEGORY:
            category_nodes.append(coord.node)

    # 计算每个 category 的"目标行数"
    cat_target_rows: Dict[int, int] = {}
    for cat_node in category_nodes:
        cat_id = cat_node.categoryId
        if cat_id is None:
            continue
        if _category_has_application_columns(cat_node):
            apps = app_cache.get(user.id, {}).get(cat_id, [])
            cat_target_rows[cat_id] = min(max(len(apps), 1), max_per_cat)
        else:
            cat_target_rows[cat_id] = 1

    # ══════ FIX: max_count 取所有"按 application 展开"的 category 的最大值 ══════
    # 只对含 application_* 子列的 category 计算展开行数；
    # user_basic / user_extra 列不参与行展开（每行重复同样的 user 值是 bug）
    max_count = max(
        (rows for rows in cat_target_rows.values() if rows > 1),
        default=1,
    )
    if max_count < 1:
        max_count = 1

    # 渲染每一行
    for r in range(max_count):
        excel_row = start_row + r
        for coord in coords:
            node = coord.node
            if not node.children:
                # application_weighted_sum 列：只写第 0 行公式
                # （其他行留空 + 行高由后续 smart height 处理）
                if node.source == APPLICATION_WEIGHTED_SUM:
                    if r == 0:
                        formula = _resolve_weighted_sum_formula(
                            col=node,
                            coords=coords,
                            start_row=start_row,
                            row_count=max_count,
                        )
                        if formula:
                            cell = ws.cell(
                                row=excel_row + 1,
                                column=coord.col_start + 1,
                                value=formula,
                            )
                            # 视觉上让单行公式在 N 行高度内居中显示
                            cell.alignment = Alignment(
                                horizontal="center", vertical="center", wrap_text=True
                            )
                    continue  # weighted_sum 列不进入普通 _resolve_leaf_value 逻辑

                # 叶子节点：写值（即使 user_basic/user_extra 在多行重复也没关系，
                # 后续 _merge_repeated_user_columns 会把相同值的连续 cell 合并）
                value = _resolve_leaf_value(
                    user=user,
                    col=node,
                    row_idx=r,
                    app_cache=app_cache,
                    extra_field_map=extra_field_map,
                    by_id=by_id,
                    max_per_cat=max_per_cat,
                )
                cell = ws.cell(
                    row=excel_row + 1,
                    column=coord.col_start + 1,
                    value=value,
                )
            else:
                # 非叶子节点：什么都不做（merge 已完成）
                pass

    # ══════ FIX: 合并 user_basic/user_extra 列在多行展开时的重复单元格 ══════
    # 当 application 列展开成 N 行时，user_basic/user_extra 列每个 user 的值在 N 行重复
    # 视觉效果差，需要把 (start_row, start_row+N-1) 这 N 个相同值的 cell 合并，
    # 且 vertical_alignment='center' 让内容垂直居中
    if max_count > 1:
        _merge_repeated_user_columns(
            ws=ws,
            start_row=start_row,
            row_count=max_count,
            coords=coords,
        )

    # ══════ application_weighted_sum 列：合并 N 行 × 1 列 → 占满展开区域 ══════
    # application_* 列在 max_count 个 application 行展开。weighted_sum 列如果只是
    # 在第 0 行写公式 + 调高行高，视觉上只有左上角一个小单元格、内容很短。
    # 这里把 weighted_sum 列在 (start_row+1, start_row+row_count) 这 N 个 cell
    # 合并成一个，让"加权公式"视觉上占满整片展开区域（与 user_basic / user_extra
    # 展开时的处理方式一致），并把行高恢复为普通单行高度（不强制放大）。
    if max_count > 1:
        for coord in coords:
            if coord.node.source != APPLICATION_WEIGHTED_SUM:
                continue
            col_idx = coord.col_start + 1  # openpyxl 1-based
            first_row = start_row + 1
            last_row = start_row + max_count
            ws.merge_cells(
                start_row=first_row,
                end_row=last_row,
                start_column=col_idx,
                end_column=col_idx,
            )
            # 合并后只在第一个 cell 有值，alignment 已是 center，写一次确保生效
            ws.cell(
                row=first_row, column=col_idx
            ).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    return start_row + max_count


def _merge_repeated_user_columns(
    ws: Worksheet,
    start_row: int,
    row_count: int,
    coords: List[NodeCoord],
) -> None:
    """合并 user_basic/user_extra 列在多行展开时的重复单元格。

    对每个 user_basic/user_extra 叶子节点：
    - 范围是 start_row+1 到 start_row+row_count（openpyxl 1-based）
    - 只有当所有 cell 值都相同时才合并
    - 合并后设置 vertical_alignment='center' 让内容垂直居中
    """  # 把 user_basic/user_extra 列在多行展开时重复的相同值垂直合并居中
    from openpyxl.styles import Alignment

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for coord in coords:
        node = coord.node
        if node.children:
            continue  # 跳过非叶子（已被表头 merge 处理）
        if node.source not in (USER_BASIC, USER_EXTRA):
            continue  # 只处理 user_basic/user_extra

        col_idx = coord.col_start + 1  # openpyxl 1-based
        first_row = start_row + 1     # openpyxl 1-based
        last_row = start_row + row_count

        # 收集所有 cell 的值
        values: List[Any] = []
        for r in range(first_row, last_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            values.append(cell.value)

        # 全部相同（且非 None）才合并
        if not values or any(v is None for v in values):
            continue
        if len(set(values)) == 1:
            # 合并范围：first_row ~ last_row
            ws.merge_cells(
                start_row=first_row,
                end_row=last_row,
                start_column=col_idx,
                end_column=col_idx,
            )
            # 让内容垂直居中
            ws.cell(row=first_row, column=col_idx).alignment = center_align


def _resolve_leaf_value(
    user: User,
    col: ExportColumnNode,
    row_idx: int,
    app_cache: Dict[int, Dict[int, List[Application]]],
    extra_field_map: Dict[str, int],
    by_id: Dict[str, ExportColumnNode],
    max_per_cat: int,
) -> Any:
    """计算单个叶子节点在指定 row 的值"""  # 根据列 source 分发到对应取值函数，得到某行某列单元格的具体值
    if col.source == USER_BASIC:
        v = _resolve_user_basic_value(user, col.basicField or "")
        if col.cellTransform == "grade":
            v = _transform_grade(v)
        return v

    if col.source == USER_EXTRA:
        if not col.fieldPath:
            return ""
        field_id = extra_field_map.get(col.fieldPath)
        if field_id is None:
            return ""
        return _resolve_user_extra_value(user, col.fieldPath, field_id)

    if col.source in CONSTRAINED_SOURCES:
        cat_id = _find_ancestor_category_id_in_map(col, by_id)
        if cat_id is None:
            return ""
        apps = app_cache.get(user.id, {}).get(cat_id, [])
        if row_idx >= len(apps):
            return ""
        app = apps[row_idx]
        return _resolve_app_value(app, col)

    return ""


# ════════════════════════════════════════════════════════════════════════
# 对外主入口
# ════════════════════════════════════════════════════════════════════════

class ExportService:
    """学生数据导出服务（v8.1）

    提供一个静态方法 stream_students_xlsx，返回 (bytes, filename)。
    调用方把 bytes 包成 StreamingResponse。
    """  # 对外暴露的导出服务入口（当前仅提供 stream_students_xlsx 一个静态方法）

    @staticmethod
    async def stream_students_xlsx(
        db: AsyncSession,
        req: ExportUsersRequest,
    ) -> Tuple[bytes, str]:
        """生成 Excel 文件字节流

        Returns:
            (xlsx_bytes, filename_with_extension)
        """  # 主流程：校验→预加载→算坐标→渲染表头→渲染数据→返回字节流和文件名
        # 1. 校验列树
        _validate_columns(req.columns)

        # 2. 收集预加载所需数据
        category_nodes = _collect_category_nodes(req.columns)
        category_ids = _collect_category_ids(category_nodes)
        field_paths = _collect_field_paths(req.columns)

        # 3. 查学生
        users = await _query_students(
            db, req.filters, req.studentIds, req.excludedIds
        )

        if not users:
            # 空表也算合法（生成只有表头的 xlsx）
            pass

        # 4. 预加载 applications + extra_field_map
        user_ids = [u.id for u in users]
        app_cache = await _preload_applications(db, user_ids, category_ids)
        extra_field_map = await _preload_extra_field_map(db, field_paths)

        # 5. 校验 extra_field_map 完整性（fieldPath 都能解析到 id）
        for fp in field_paths:
            if fp not in extra_field_map:
                raise BadRequestError(
                    f"扩展信息字段名 '{fp}' 在 extra_info_field 表中不存在"
                )

        # 6. 分配坐标 + 建 by_id_map
        coords = _assign_coordinates(req.columns)
        by_id = _build_by_id_map(req.columns)

        # 7. 创建 workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "学生数据"

        # 8. 渲染表头
        header_row_count = _render_header(ws, coords)

        # 9. 列宽自适应（按 label 字符数粗算）
        for coord in coords:
            col_letter = get_column_letter(coord.col_start + 1)
            # 中文 1 字 = 2 英文字符宽
            label_width = sum(2 if ord(ch) > 127 else 1 for ch in coord.node.label)
            ws.column_dimensions[col_letter].width = max(label_width + 4, 12)

        # 10. 渲染数据
        max_per_cat = req.maxApplicationsPerCategory
        # ══════ FIX: start_row 必须推进 ══════
        # 原 bug: 每个 user 都从 header_row_count 开始写，后一个 user 覆盖前一个
        # 修复: 累积 current_row，每次 _render_data 后推进
        current_row = header_row_count
        for user in users:
            current_row = _render_data(
                ws=ws,
                start_row=current_row,
                user=user,
                coords=coords,
                app_cache=app_cache,
                extra_field_map=extra_field_map,
                by_id=by_id,
                max_per_cat=max_per_cat,
            )

        # 11. 冻结表头
        if header_row_count > 0:
            ws.freeze_panes = ws.cell(row=header_row_count + 1, column=1)

        # ══════ DEBUG: 打印关键信息 ══════
        logger.info(
            f"[export-debug] users.count={len(users)}, "
            f"columns={len(req.columns)}, "
            f"coords.count={len(coords)}, "
            f"header_row_count={header_row_count}, "
            f"max_per_cat={req.maxApplicationsPerCategory}"
        )
        if users:
            sample = users[0]
            logger.info(
                f"[export-debug] sample user: id={sample.id}, username={sample.username}, "
                f"major={sample.major!r}, department={sample.department!r}, "
                f"extra_info_keys={list((sample.extra_info or {}).keys())[:5]}"
            )

        # ══════ DEBUG: 打印每个 leaf 列的写入位置 ══════
        leaf_coords = [(c.col_start, c.node.label, c.node.source, c.node.fieldPath) for c in coords if not c.node.children]
        logger.info(f"[export-debug] leaf coords ({len(leaf_coords)}): {leaf_coords}")

        # 12. 写入 bytes
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx_bytes = buf.getvalue()

        # 13. 文件名（sanitize 一下，避免非法字符）
        safe_name = _sanitize_filename(req.fileName)
        date_suffix = _today_yyyymmdd()
        filename = f"{safe_name}-{date_suffix}.xlsx"

        return xlsx_bytes, filename


def _sanitize_filename(name: str) -> str:
    """清洗文件名（去路径分隔符、特殊字符）"""  # 清洗非法文件名字符并截断长度，保证安全文件名
    # Windows 不允许的字符: < > : " / \ | ? *
    bad_chars = '<>:"/\\|?*'
    result = "".join("_" if c in bad_chars else c for c in name).strip()
    if not result:
        result = "students"
    return result[:50]  # 截断


def _today_yyyymmdd() -> str:
    """生成当前日期的 YYYYMMDD 字符串，用于文件名后缀"""  # 取当前日期的 YYYYMMDD 字符串
    from datetime import datetime
    return datetime.now().strftime("%Y%m%d")


__all__ = ["ExportService"]
