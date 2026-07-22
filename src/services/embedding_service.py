"""Embedding 服务（Layer 2）

职责：
|- 文件解析（PDF/DOCX/XLSX/TXT）
|- 文本切块
|- 向量 upsert / search
|- CRUD 业务方法（管理端）

注意：LLM/Embedding 模型调用已迁移到 infra.llm
"""
import io
import json
import logging
from typing import List, Optional, Tuple

from sqlalchemy.ext.asyncio import AsyncSession

from src.infra.llm import get_embedding_model, embed_text, embed_texts
from src.infra.config import get_settings
from src.models.embedding import EmbeddingCategory
from src.repositories.embedding_repo import EmbeddingRepository
from src.app.schemas.embedding import (
    EmbeddingUploadRequest,
    EmbeddingUpdateRequest,
    EmbeddingQueryRequest,
    EmbeddingDeleteRequest,
    EmbeddingSearchRequest,
    EmbeddingVO,
    EmbeddingDetailVO,
    EmbeddingSearchResultVO,
    EmbeddingUploadResultVO,
    EmbeddingDeleteResultVO,
    EmbeddingStatsVO,
    EmbeddingListVO,
    EmbeddingSearchListVO,
    Page,
)

logger = logging.getLogger(__name__)


class EmbeddingService:
    """Embedding 服务（Layer 2）"""

    def __init__(self):
        self.settings = get_settings()

    # ═══════════════════════════════════════════════════════════════════════════════
    # 文件解析（PDF/DOCX/XLSX/TXT）
    # ═══════════════════════════════════════════════════════════════════════════════

    @staticmethod
    def parse_file(file_bytes: bytes, filename: str) -> str:
        """解析文件为文本。

        Args:
            file_bytes: 文件字节内容
            filename: 文件名（用于判断类型）

        Returns:
            提取的文本内容
        """
        ext = filename.lower().split(".")[-1] if "." in filename else ""

        parsers = {
            "pdf": EmbeddingService._parse_pdf,
            "docx": EmbeddingService._parse_docx,
            "doc": EmbeddingService._parse_docx,
            "xlsx": EmbeddingService._parse_xlsx,
            "xls": EmbeddingService._parse_xlsx,
            "txt": EmbeddingService._parse_txt,
        }

        parser = parsers.get(ext)
        if not parser:
            logger.warning(f"不支持的文件类型: {ext}, 文件名: {filename}")
            return ""

        try:
            return parser(file_bytes)
        except Exception as e:
            logger.error(f"解析文件失败: {filename}, 错误: {e}")
            return ""

    @staticmethod
    def _parse_pdf(file_bytes: bytes) -> str:
        """解析 PDF 文件。"""
        try:
            import pdfplumber

            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                texts = []
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        texts.append(text)
                return "\n".join(texts)
        except ImportError:
            logger.warning("pdfplumber 未安装，使用 PyPDF2 作为备选")
            return EmbeddingService._parse_pdf_pypdf2(file_bytes)
        except Exception as e:
            logger.error(f"PDF 解析失败: {e}")
            return ""

    @staticmethod
    def _parse_pdf_pypdf2(file_bytes: bytes) -> str:
        """使用 PyPDF2 解析 PDF（备选方案）。"""
        try:
            from PyPDF2 import PdfReader

            reader = PdfReader(io.BytesIO(file_bytes))
            texts = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    texts.append(text)
            return "\n".join(texts)
        except ImportError:
            logger.error("PyPDF2 也未安装，无法解析 PDF")
            return ""
        except Exception as e:
            logger.error(f"PyPDF2 解析失败: {e}")
            return ""

    @staticmethod
    def _parse_docx(file_bytes: bytes) -> str:
        """解析 Word 文件（.docx）。"""
        try:
            import mammoth

            result = mammoth.extract_raw_text(io.BytesIO(file_bytes))
            if result.messages:
                for msg in result.messages:
                    logger.debug(f"mammoth 消息: {msg}")
            return result.value
        except ImportError:
            logger.error("mammoth 未安装，无法解析 docx")
            return ""
        except Exception as e:
            logger.error(f"DOCX 解析失败: {e}")
            return ""

    @staticmethod
    def _parse_xlsx(file_bytes: bytes) -> str:
        """解析 Excel 文件（.xlsx/.xls）。"""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            texts = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                texts.append(f"[Sheet: {sheet_name}]")

                for row in sheet.iter_rows(values_only=True):
                    row_values = [str(cell) for cell in row if cell is not None]
                    if row_values:
                        texts.append(" | ".join(row_values))

            return "\n".join(texts)
        except ImportError:
            logger.error("openpyxl 未安装，无法解析 xlsx")
            return ""
        except Exception as e:
            logger.error(f"XLSX 解析失败: {e}")
            return ""

    @staticmethod
    def _parse_txt(file_bytes: bytes) -> str:
        """解析纯文本文件。"""
        try:
            return file_bytes.decode("utf-8", errors="ignore")
        except Exception as e:
            logger.error(f"TXT 解析失败: {e}")
            return ""

    # ---------- 核心：生成 embedding ----------

    async def generate_embedding(self, text: str) -> List[float]:
        """调用 Embedding API，生成文本向量。

        Args:
            text: 输入文本（建议不超过 8192 tokens）

        Returns:
            向量列表，长度为 EMBEDDING_DIM
        """
        if not text or not text.strip():
            raise ValueError("text 不能为空")

        return await embed_text(text)

    async def generate_embeddings_batch(
        self, texts: List[str]
    ) -> List[List[float]]:
        """批量生成 embedding（API 支持单次多文本）。"""
        if not texts:
            return []

        return await embed_texts(texts)

    # ---------- 文本切块 ----------

    @staticmethod
    def chunk_text(
        text: str,
        *,
        chunk_size: Optional[int] = None,
        overlap: Optional[int] = None,
    ) -> List[str]:
        """按字符数切块（简单实现，可替换为语义切块）。

        Args:
            text: 原始文本
            chunk_size: 每块字符数，默认 RAG_CHUNK_SIZE
            overlap: 块之间重叠字符数

        Returns:
            切块后的文本列表
        """
        settings = get_settings()
        size = chunk_size or settings.RAG_CHUNK_SIZE
        overlap = overlap or settings.RAG_CHUNK_OVERLAP

        if len(text) <= size:
            return [text]

        chunks = []
        start = 0
        while start < len(text):
            end = start + size
            chunks.append(text[start:end])
            start += size - overlap

        return chunks

    # ---------- CRUD：upsert / delete / search ----------

    async def upsert(
        self,
        db: AsyncSession,
        *,
        title: Optional[str],
        content: str,
        category: str,
        ref_id: Optional[int] = None,
    ) -> Tuple[EmbeddingRepository, bool]:
        """解析文本 → 切块 → 生成向量 → 存入数据库。

        Args:
            db: 数据库会话
            title: 标题（方便人类识别）
            content: 内容原文
            category: EmbeddingCategory 枚举值
            ref_id: 关联业务 ID（如 template.id）

        Returns:
            (embedding, is_new) - 新增返回 True，更新返回 False
        """
        chunks = self.chunk_text(content)
        if not chunks:
            raise ValueError(f"内容为空或解析失败: {title}")

        # 取第一个 chunk 生成向量（完整内容可能超长，可优化为多 chunk 拼接）
        vector = await self.generate_embedding(chunks[0])

        embedding = await EmbeddingRepository.upsert(
            db,
            title=title,
            content=content,
            category=category,
            ref_id=ref_id,
            embedding_vector=vector,
        )
        await EmbeddingRepository.commit(db)

        is_new = embedding.id is not None
        return embedding, is_new

    async def upsert_by_ref(
        self,
        db: AsyncSession,
        *,
        title: Optional[str],
        content: str,
        category: EmbeddingCategory,
        ref_id: int,
    ) -> None:
        """按 ref_id 更新向量（用于模板更新时同步）。

        category + ref_id 唯一，upsert 语义。
        """
        await self.upsert(
            db,
            title=title,
            content=content,
            category=category.value,
            ref_id=ref_id,
        )

    async def delete_by_ref(
        self,
        db: AsyncSession,
        category: EmbeddingCategory,
        ref_id: int,
    ) -> int:
        """按 ref_id 删除向量（用于模板删除时同步）。

        Returns:
            删除的行数
        """
        count = await EmbeddingRepository.delete_by_ref_id(
            db,
            category=category.value,
            ref_id=ref_id,
        )
        await EmbeddingRepository.commit(db)
        return count

    # ---------- 向量搜索 ----------

    async def search(
        self,
        db: AsyncSession,
        query: str,
        *,
        category: Optional[str] = None,
        top_k: int = 5,
    ) -> List[dict]:
        """向量相似度搜索。

        Args:
            db: 数据库会话
            query: 用户查询文本
            category: 可选，限制 category
            top_k: 返回数量

        Returns:
            [{"id", "title", "content", "score"}, ...]
        """
        # 1. 生成 query 向量
        query_vector = await self.generate_embedding(query)

        # 2. 从 DB 加载候选向量（可加 category 过滤）
        if category:
            candidates = await EmbeddingRepository.list_by_category(
                db, category, limit=1000
            )
        else:
            # TODO: 实现全量搜索（分页加载）
            candidates = await EmbeddingRepository.list_by_category(
                db, EmbeddingCategory.POLICY.value, limit=1000
            )

        # 3. 计算余弦相似度并排序
        scored = []
        for emb in candidates:
            if emb.embedding is None:
                continue
            score = self._cosine_similarity(query_vector, emb.embedding)
            scored.append({
                "id": emb.id,
                "title": emb.title,
                "content": emb.content,
                "category": emb.category,
                "ref_id": emb.ref_id,
                "score": score,
            })

        # 4. 排序取 top_k
        scored.sort(key=lambda x: x["score"], reverse=True)
        return scored[:top_k]

    @staticmethod
    def _cosine_similarity(a: List[float], b: List[float]) -> float:
        """计算余弦相似度。"""
        if len(a) != len(b):
            return 0.0

        dot = sum(x * y for x, y in zip(a, b))
        norm_a = sum(x * x for x in a) ** 0.5
        norm_b = sum(x * x for x in b) ** 0.5

        if norm_a == 0 or norm_b == 0:
            return 0.0

        return dot / (norm_a * norm_b)

    # ═══════════════════════════════════════════════════════════════════════════
    # 管理端 API（上传/删除/查询）
    # ═══════════════════════════════════════════════════════════════════════════

    async def upload(
        self,
        db: AsyncSession,
        request: EmbeddingUploadRequest,
    ) -> EmbeddingUploadResultVO:
        """上传并索引 embedding。

        流程：文本 → 切块 → 生成向量 → 存入数据库
        """
        # 校验 category
        request.validate_category()

        # 生成向量并存储
        embedding, is_new = await self.upsert(
            db,
            title=request.title,
            content=request.content,
            category=request.category,
            ref_id=None,  # 管理端手动上传不关联业务
        )

        return EmbeddingUploadResultVO.from_orm_to_vo(embedding, is_new)

    async def update(
        self,
        db: AsyncSession,
        embedding_id: int,
        request: EmbeddingUpdateRequest,
    ) -> Optional[EmbeddingVO]:
        """更新 embedding（仅更新文本，向量需要重新生成）"""
        embedding = await EmbeddingRepository.get_by_id(db, embedding_id)
        if not embedding:
            return None

        # 更新文本字段
        if request.title is not None:
            embedding.title = request.title
        if request.content is not None:
            embedding.content = request.content
            # 重新生成向量
            vector = await self.generate_embedding(request.content)
            embedding.embedding = vector

        await EmbeddingRepository.update(db, embedding)
        await EmbeddingRepository.commit(db)

        return EmbeddingVO.from_orm_to_vo(embedding)

    async def delete(
        self,
        db: AsyncSession,
        request: EmbeddingDeleteRequest,
    ) -> EmbeddingDeleteResultVO:
        """批量删除 embedding"""
        deleted_count = await EmbeddingRepository.delete_by_ids(db, request.ids)
        await EmbeddingRepository.commit(db)

        return EmbeddingDeleteResultVO(
            deletedCount=deleted_count,
            totalRequested=len(request.ids),
        )

    async def list_(
        self,
        db: AsyncSession,
        request: EmbeddingQueryRequest,
    ) -> EmbeddingListVO:
        """分页查询 embedding 列表"""
        items, total = await EmbeddingRepository.paginate(
            db,
            category=request.category,
            keyword=request.keyword,
            page_num=request.page_num,
            page_size=request.page_size,
        )

        vos = [EmbeddingVO.from_orm_to_vo(emb) for emb in items]
        return Page.from_list_to_page(vos, total, request.page_num, request.page_size)

    async def get_detail(
        self,
        db: AsyncSession,
        embedding_id: int,
    ) -> Optional[EmbeddingDetailVO]:
        """获取 embedding 详情（含向量）"""
        embedding = await EmbeddingRepository.get_by_id(db, embedding_id)
        if not embedding:
            return None
        return EmbeddingDetailVO.from_orm_to_vo(embedding, include_vector=True)

    async def search_(
        self,
        db: AsyncSession,
        request: EmbeddingSearchRequest,
    ) -> EmbeddingSearchListVO:
        """向量语义搜索"""
        results = await self.search(
            db,
            query=request.query,
            category=request.category,
            top_k=request.top_k,
        )

        vos = [EmbeddingSearchResultVO.from_search_result(r) for r in results]
        total = len(vos)
        return Page.from_list_to_page(vos, total, 1, request.top_k)

    async def get_stats(self, db: AsyncSession) -> EmbeddingStatsVO:
        """获取统计信息"""
        total = await EmbeddingRepository.count_all(db)
        category_stats = await EmbeddingRepository.get_category_stats(db)

        return EmbeddingStatsVO(
            totalCount=total,
            categoryStats=category_stats,
        )


# 全局单例（按需初始化）
_embedding_service: Optional[EmbeddingService] = None


def get_embedding_service() -> EmbeddingService:
    """获取 EmbeddingService 单例。"""
    global _embedding_service
    if _embedding_service is None:
        _embedding_service = EmbeddingService()
    return _embedding_service
